import re
from typing import List, Dict, Any
import io

try:
    import openpyxl

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class XLSXParser:
    SUPPORTED_EXTENSIONS = [".xlsx", ".xls"]

    @staticmethod
    def parse(file_bytes: bytes) -> Dict[str, Any]:
        urls = []
        text = ""
        metadata = {}
        scripts = []
        macros_detected = False

        if not OPENPYXL_AVAILABLE:
            text = XLSXParser._extract_text_raw(file_bytes)
            urls = XLSXParser._extract_urls_from_text(text)
            return {
                "text": text,
                "urls": urls,
                "metadata": metadata,
                "scripts": scripts,
                "macros_detected": macros_detected,
            }

        try:
            with io.BytesIO(file_bytes) as f:
                wb = openpyxl.load_workbook(f, data_only=True)

                for sheet in wb:
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value:
                                text += str(cell.value) + " "

                urls.extend(XLSXParser._extract_urls_from_text(text))
                urls = list(set(urls))

                metadata = {
                    "sheet_count": len(wb.sheetnames),
                    "sheets": wb.sheetnames,
                }

                macros_detected = XLSXParser._check_for_macros(file_bytes)

        except Exception as e:
            text = XLSXParser._extract_text_raw(file_bytes)
            urls = XLSXParser._extract_urls_from_text(text)

        return {
            "text": text,
            "urls": urls,
            "metadata": metadata,
            "scripts": scripts,
            "macros_detected": macros_detected,
        }

    @staticmethod
    def _extract_text_raw(file_bytes: bytes) -> str:
        try:
            return file_bytes.decode("utf-8", errors="ignore")
        except:
            return ""

    @staticmethod
    def _extract_urls_from_text(text: str) -> List[str]:
        url_pattern = re.compile(
            r"https?://(?:[-\w.]|(?:%[\da-fA-F]{2}))+[^\s]*", re.IGNORECASE
        )
        return url_pattern.findall(text)

    @staticmethod
    def _check_for_macros(file_bytes: bytes) -> bool:
        macro_indicators = [b"vbaProject.bin", b"VBA", b"xl macros", b"bin VBA"]
        return any(indicator in file_bytes for indicator in macro_indicators)
